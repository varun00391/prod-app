import base64
import io
import os
import tempfile
from groq import Groq
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from docx import Document

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}
ALLOWED_AUDIO_TYPES = {"audio/mpeg", "audio/mp3", "audio/wav", "audio/webm", "audio/m4a", "audio/mp4"}
WHISPER_MODEL = "whisper-large-v3-turbo"

# Groq models: text-only vs vision (when request contains images)
# Vision: https://console.groq.com/docs/vision
MODEL_TEXT = "llama-3.3-70b-versatile"
MODEL_VISION = "meta-llama/llama-4-scout-17b-16e-instruct"


def extract_text_from_pdf(file_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(file_bytes))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def extract_text_from_excel(file_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            parts.append("\t".join(str(c) if c is not None else "" for c in row))
    return "\n".join(parts)


def extract_text_from_docx(file_bytes: bytes) -> str:
    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs)


def _transcribe_audio(audio_bytes: bytes, client: Groq, filename: str = "audio") -> str:
    """Transcribe audio using Groq Whisper. Returns transcript text or raises."""
    ext = ".mp3"  # Groq accepts mp3, wav, m4a, webm, etc.
    if filename:
        for e in (".wav", ".m4a", ".webm", ".mp3", ".mpeg", ".mpga", ".ogg", ".flac"):
            if filename.lower().endswith(e):
                ext = e
                break
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
        f.write(audio_bytes)
        f.flush()
        try:
            with open(f.name, "rb") as audio_file:
                resp = client.audio.transcriptions.create(
                    file=audio_file,
                    model=WHISPER_MODEL,
                    response_format="text",
                )
            return (resp.text if hasattr(resp, "text") else resp) or ""
        finally:
            os.unlink(f.name)


def build_messages(messages: list[dict], api_key: str | None) -> tuple[list[dict], bool]:
    """Build chat messages for Groq. Returns (messages, has_images)."""
    key = api_key or os.environ.get("GROQ_API_KEY")
    if not key:
        raise ValueError("Groq API key is required (header X-API-Key or env GROQ_API_KEY)")
    client = Groq(api_key=key)

    out = []
    has_images = False
    for m in messages:
        role = m.get("role", "user")
        content = m.get("content")
        attachments = m.get("attachments") or []

        parts = []
        if isinstance(content, str) and content.strip():
            parts.append({"type": "text", "text": content})

        for att in attachments:
            name = att.get("name", "")
            data = att.get("data")
            mime_type = (att.get("mime_type") or "").lower()
            url = att.get("url")

            if url and not data:
                if mime_type in ALLOWED_IMAGE_TYPES:
                    parts.append({"type": "image_url", "image_url": {"url": url}})
                    has_images = True
                else:
                    parts.append({"type": "text", "text": f"[Attachment: {name}]\n(URL provided)"})
                continue

            if data:
                b64 = data if isinstance(data, str) else base64.b64encode(data).decode()
                if mime_type in ALLOWED_IMAGE_TYPES:
                    parts.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:{mime_type};base64,{b64}"}
                    })
                    has_images = True
                elif mime_type in ALLOWED_AUDIO_TYPES:
                    raw = base64.b64decode(b64)
                    try:
                        transcript = _transcribe_audio(raw, client, name)
                        parts.append({
                            "type": "text",
                            "text": f"[Audio transcript from {name}]:\n\n{transcript}"
                        })
                    except Exception:
                        parts.append({
                            "type": "text",
                            "text": f"[Audio: {name}] Could not transcribe this file. Please describe the content or paste a transcript."
                        })
                elif "pdf" in mime_type:
                    raw = base64.b64decode(b64)
                    text = extract_text_from_pdf(raw)
                    parts.append({"type": "text", "text": f"[PDF: {name}]\n\n{text}"})
                elif "spreadsheet" in mime_type or "excel" in mime_type or mime_type in (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                ):
                    raw = base64.b64decode(b64)
                    text = extract_text_from_excel(raw)
                    parts.append({"type": "text", "text": f"[Excel: {name}]\n\n{text}"})
                elif "wordprocessingml" in mime_type or "msword" in mime_type:
                    raw = base64.b64decode(b64)
                    text = extract_text_from_docx(raw)
                    parts.append({"type": "text", "text": f"[Word: {name}]\n\n{text}"})
                else:
                    parts.append({"type": "text", "text": f"[Attachment: {name}] (binary content not extracted)"})

        if not parts:
            continue
        if len(parts) == 1 and parts[0]["type"] == "text":
            out.append({"role": role, "content": parts[0]["text"]})
        else:
            out.append({"role": role, "content": parts})
    return out, has_images


def chat_completion(messages: list[dict], api_key: str | None) -> str:
    key = api_key or os.environ.get("GROQ_API_KEY")
    if not key:
        raise ValueError("Groq API key is required (header X-API-Key or env GROQ_API_KEY)")

    client = Groq(api_key=key)
    groq_messages, has_images = build_messages(messages, api_key)
    model = MODEL_VISION if has_images else MODEL_TEXT

    resp = client.chat.completions.create(
        model=model,
        messages=groq_messages,
        max_tokens=4096,
    )
    return resp.choices[0].message.content or ""
